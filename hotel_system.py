import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# ================================================
# YOUR DETAILS - CHANGE THESE
# ===============================================
from dotenv import load_dotenv
import os

load_dotenv()
sender_email = os.getenv("SENDER_EMAIL")
app_password = os.getenv("APP_PASSWORD")

# =================================================
# STEP 1 - CREATE GUEST DATA
# =================================================
workbook = openpyxl.Workbook()
sheet = workbook.active

sheet["A1"] = "Name"
sheet["B1"] = "Email"
sheet["C1"] = "Room"
sheet["D1"] = "Nights"
sheet["E1"] = "Price Per Night"
sheet["F1"] = "Total Bill"

guests = [
    ("Charan", sender_email, 101, 3, 2000),
    ("Monica", sender_email, 102, 2, 2500),
    ("Raj", sender_email, 103, 5, 1500),
]
row = 2
for guest in guests:
    name, email, room, nights, price = guest
    total = nights * price
    sheet.cell(row=row, column=1, value=name)
    sheet.cell(row=row, column=2, value=email)
    sheet.cell(row=row, column=3, value=room)
    sheet.cell(row=row, column=4, value=nights)
    sheet.cell(row=row, column=5, value=price)
    sheet.cell(row=row, column=6, value=total)
    row += 1

workbook.save("c:/Users/SK MUJEEB/Desktop/hotel_guests.xlsx")
print("Guest Excel file created!")

# ====================================
# STEP 2 - SEND WELCOME EMAILS
# ====================================
def send_email(name, receiver, room, nights, total):
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver
    message["Subject"] = "Welcome to The Golkonda Hotel, " + name + "!"

    body = """Dear """ + name + """,

Welcome to The Golkonda Hotel! We are delighted to have you!

Your booking details:
- Room Number: """ + str(room) + """
- Number of Nights: """ + str(nights) + """
- Total Bill: Rs """ + str(total) + """

We hope you have a wonderful stay!

Best regards,
Hotel Management"""

    message.attach(MIMEText(body, "plain"))

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender_email, app_password)
        server.sendmail(sender_email, receiver, message.as_string())
        server.quit()
        print("Email sent to", name)
    except:
        print("Failed to send email to", name)

# Send email to all guests              
for guest in guests:
    name, email, room, nights, price = guest
    total = nights * price
    send_email(name, email, room, nights, total)

# ================================================    
# STEP 3 - GENERATE HOTEL REPORT
# ================================================
report = openpyxl.Workbook()
sheet2 = report.active

sheet2["A1"] = "HOTEL DAILY REPORT "
sheet2["A3"] = "Total Guests"
sheet2["A4"] = "Total Rooms Occupied"
sheet2["A5"] = "Total Revenue"
sheet2["A6"] = "Average Bill Per Guest"

total_guests = len(guests)
total_revenue = sum(n * p for _, _, _, n, p in guests)
average_bill = total_revenue // total_guests

sheet2["B3"] = total_guests
sheet2["B4"] = total_guests
sheet2["B5"] = "Rs " + str(total_revenue)
sheet2["B6"] = "Rs " + str(average_bill)

report.save("c:/Users/SK MUJEEB/Desktop/hotel_report.xlsx")
print("Hotel report generated!")
print("---")
print("SUMMARY:")
print("Total Guests:", total_guests)
print("Total Revenue: Rs", total_revenue)
print("Average Bill: Rs", average_bill)

